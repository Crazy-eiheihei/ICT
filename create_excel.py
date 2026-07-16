import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Dữ liệu các thông số cáp quang
data = [
    ["STT", "Tên Thông Số", "Giá Trị", "Đơn Vị", "Ghi Chú"],
    [1, "Đường kính lõi", "9-10", "μm", "Cáp quang đơn mode"],
    [2, "Đường kính áo", "125", "μm", "Đã được chuẩn hóa"],
    [3, "Aperture Số (NA)", "0.13-0.14", "", "Xác định góc thu nhận ánh sáng"],
    [4, "Suy hao (Attenuation)", "0.2-0.35", "dB/km", "Tại bước sóng 1550nm"],
    [5, "Phân tán (Dispersion)", "16-17", "ps/nm/km", "Tại bước sóng 1550nm"],
    [6, "Bước sóng hoạt động", "1310-1550", "nm", "Phạm vi hoạt động chính"],
    [7, "Độ uốn tối thiểu", "30", "mm", "Tính từ tâm lõi"],
    [8, "Độ bền kéo tối đa", "200", "MPa", "Lực kéo tối đa"],
    [9, "Nhiệt độ hoạt động", "-40 đến 70", "°C", "Phạm vi nhiệt độ"],
    [10, "Loại cáp", "Single Mode (SM)", "", "Cáp quang đơn mode hay đa mode"],
    [11, "Chuẩn", "ITU-T G.652", "", "Chuẩn quốc tế"],
    [12, "Độ dài cáp", "Tuỳ chỉnh", "km", "Tùy theo yêu cầu ứng dụng"],
    [13, "Đầu nối", "FC/APC hoặc LC/APC", "", "Các loại đầu nối phổ biến"],
    [14, "Cấu trúc áo bảo vệ", "Nylon hoặc Steel", "", "Bảo vệ cáp quang"],
    [15, "Tốc độ truyền dữ liệu", "Có thể lên đến Tbps", "bit/s", "Tùy theo thiết bị"],
]

# Tạo workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Thông số Cáp Quang"

# Thêm dữ liệu
for row_idx, row_data in enumerate(data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)

# Định dạng header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx in range(1, 6):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Định dạng các ô dữ liệu
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row_idx in range(1, len(data) + 1):
    for col_idx in range(1, 6):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Điều chỉnh độ rộng cột
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 35

# Điều chỉnh độ cao hàng
ws.row_dimensions[1].height = 30
for row_idx in range(2, len(data) + 1):
    ws.row_dimensions[row_idx].height = 25

# Lưu file
wb.save("Các thông số chính của cáp quang.xlsx")
print("File Excel đã được tạo thành công: Các thông số chính của cáp quang.xlsx")
